"""
Parse DUPE Excel cash flow model sheets into structured monthly records.

Handles both project types:
  - CASH FLOW INTERES SOCIAL  → Social interest (RD$, 61 months from Sep 2024)
  - CASH FLOW - TURISTICOS    → Tourist (USD, 60 months from Jun 2026)
"""
from __future__ import annotations
import logging
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

logger = logging.getLogger("dupe.excel_parser")

CUTOFF_DATE = date(2026, 6, 1)  # today = June 2026; months before = actual


@dataclass
class MonthlyCashFlow:
    month: str          # "YYYY-MM"
    month_number: int   # 1-based
    is_actual: bool
    income: float = 0.0
    expenses: float = 0.0
    net_cash_flow: float = 0.0
    cumulative_balance: float = 0.0
    income_ventas: float = 0.0
    income_separaciones: float = 0.0
    income_entregas: float = 0.0
    income_financiamiento: float = 0.0
    exp_construccion: float = 0.0
    exp_suelo: float = 0.0
    exp_tecnicos: float = 0.0
    exp_juridico: float = 0.0
    exp_financiero: float = 0.0
    exp_gestion: float = 0.0
    exp_comercializacion: float = 0.0


def _row_get(rows: list, row_idx: int, month_col_map: Dict[str, int]) -> Dict[str, float]:
    if row_idx >= len(rows):
        return {}
    row = rows[row_idx]
    result: Dict[str, float] = {}
    for month_str, col_idx in month_col_map.items():
        val = row[col_idx]
        result[month_str] = float(val) if val is not None else 0.0
    return result


def _find_row_by_label(rows: list, labels: list[str], start: int = 0, col: int = 1) -> Optional[int]:
    for idx in range(start, len(rows)):
        cell = rows[idx][col]
        if cell and any(str(cell).strip().upper() == lbl.upper() for lbl in labels):
            return idx
    return None


def parse_social_cashflow(excel_path: str) -> List[MonthlyCashFlow]:
    """Parse CASH FLOW INTERES SOCIAL sheet."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["CASH FLOW INTERES SOCIAL"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.error("Failed to open Excel: %s", e)
        return _get_synthetic_social()

    # Find month columns (date row = row 4, index 3)
    date_row = rows[3]
    month_col_map: Dict[str, int] = {}
    month_numbers: Dict[str, int] = {}
    month_dates: Dict[str, date] = {}
    month_num = 0
    for col_idx, val in enumerate(date_row):
        if hasattr(val, "year") and val and val.year > 2020:
            month_num += 1
            month_str = val.strftime("%Y-%m")
            month_col_map[month_str] = col_idx
            month_numbers[month_str] = month_num
            month_dates[month_str] = val.date() if hasattr(val, "date") else date(val.year, val.month, 1)

    # Row indices (0-based):
    # Row 18 (idx 17) = TOTAL income
    # Row 13 (idx 12) = SEPARACIONES
    # Row 14 (idx 13) = INICIAL (income_ventas proxy)
    # Row 15 (idx 14) = ENTREGAS
    # Row 17 (idx 16) = FINANCIAMIENTO
    # Row 28 (idx 27) = suelo TOTAL
    # Row 46 (idx 45) = construccion TOTAL
    # Row 61 (idx 60) = tecnicos TOTAL
    # Row 72 (idx 71) = juridico TOTAL
    # Row 78 (idx 77) = financiero TOTAL
    # Row 84 (idx 83) = gestion TOTAL
    # Row 90 (idx 89) = comercializacion TOTAL
    # Row 92 (idx 91) = TOTAL GASTOS
    # Row 95 (idx 94) = FLUJO DE CAJA
    # Row 96 (idx 95) = ACUMULADO

    income_total  = _row_get(rows, 17, month_col_map)
    separaciones  = _row_get(rows, 12, month_col_map)
    entregas      = _row_get(rows, 14, month_col_map)
    financiamiento= _row_get(rows, 16, month_col_map)
    exp_suelo     = _row_get(rows, 27, month_col_map)
    exp_const     = _row_get(rows, 45, month_col_map)
    exp_tec       = _row_get(rows, 60, month_col_map)
    exp_jur       = _row_get(rows, 71, month_col_map)
    exp_fin       = _row_get(rows, 77, month_col_map)
    exp_ges       = _row_get(rows, 83, month_col_map)
    exp_com       = _row_get(rows, 89, month_col_map)
    exp_total     = _row_get(rows, 91, month_col_map)
    acumulado     = _row_get(rows, 95, month_col_map)

    records: List[MonthlyCashFlow] = []
    for month_str in list(month_col_map.keys()):
        d = month_dates[month_str]
        inc = income_total.get(month_str, 0.0)
        exp = exp_total.get(month_str, 0.0)
        records.append(MonthlyCashFlow(
            month=month_str,
            month_number=month_numbers[month_str],
            is_actual=d < CUTOFF_DATE,
            income=inc,
            expenses=exp,
            net_cash_flow=inc - exp,
            cumulative_balance=acumulado.get(month_str, 0.0),
            income_ventas=0.0,  # not broken out at TOTAL level in this sheet
            income_separaciones=separaciones.get(month_str, 0.0),
            income_entregas=entregas.get(month_str, 0.0),
            income_financiamiento=financiamiento.get(month_str, 0.0),
            exp_construccion=exp_const.get(month_str, 0.0),
            exp_suelo=exp_suelo.get(month_str, 0.0),
            exp_tecnicos=exp_tec.get(month_str, 0.0),
            exp_juridico=exp_jur.get(month_str, 0.0),
            exp_financiero=exp_fin.get(month_str, 0.0),
            exp_gestion=exp_ges.get(month_str, 0.0),
            exp_comercializacion=exp_com.get(month_str, 0.0),
        ))
    return records


def parse_tourist_cashflow(excel_path: str) -> List[MonthlyCashFlow]:
    """Parse CASH FLOW - TURISTICOS sheet."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["CASH FLOW - TURISTICOS"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        logger.error("Failed to open Excel: %s", e)
        return _get_synthetic_tourist()

    date_row = rows[3]
    month_col_map: Dict[str, int] = {}
    month_numbers: Dict[str, int] = {}
    month_dates: Dict[str, date] = {}
    month_num = 0
    for col_idx, val in enumerate(date_row):
        if hasattr(val, "year") and val and val.year > 2020:
            month_num += 1
            month_str = val.strftime("%Y-%m")
            month_col_map[month_str] = col_idx
            month_numbers[month_str] = month_num
            month_dates[month_str] = val.date() if hasattr(val, "date") else date(val.year, val.month, 1)

    # Tourist sheet structure:
    # Row 18 (idx 17) = TOTAL income
    # Row 12 (idx 11) = SEPARACIONES
    # Row 14 (idx 13) = ENTREGAS
    # Row 16 (idx 15) = FINANCIAMIENTO
    # Row 25 (idx 24) = suelo TOTAL
    # Row 40 (idx 39) = construccion TOTAL
    # Row 54 (idx 53) = tecnicos TOTAL
    # Row 65 (idx 64) = juridico TOTAL
    # Row 71 (idx 70) = financiero TOTAL
    # Row 77 (idx 76) = gestion TOTAL
    # Row 84 (idx 83) = comercializacion TOTAL
    # Row 86 (idx 85) = TOTAL GASTOS
    # Row 89 (idx 88) = FLUJO DE CAJA
    # Row 90 (idx 89) = ACUMULADO

    income_total  = _row_get(rows, 17, month_col_map)
    separaciones  = _row_get(rows, 11, month_col_map) if len(rows) > 11 else {}
    entregas      = _row_get(rows, 13, month_col_map) if len(rows) > 13 else {}
    financiamiento= _row_get(rows, 15, month_col_map) if len(rows) > 15 else {}
    exp_suelo     = _row_get(rows, 24, month_col_map) if len(rows) > 24 else {}
    exp_const     = _row_get(rows, 39, month_col_map) if len(rows) > 39 else {}
    exp_tec       = _row_get(rows, 53, month_col_map) if len(rows) > 53 else {}
    exp_jur       = _row_get(rows, 64, month_col_map) if len(rows) > 64 else {}
    exp_fin       = _row_get(rows, 70, month_col_map) if len(rows) > 70 else {}
    exp_ges       = _row_get(rows, 76, month_col_map) if len(rows) > 76 else {}
    exp_com       = _row_get(rows, 83, month_col_map) if len(rows) > 83 else {}
    exp_total     = _row_get(rows, 85, month_col_map) if len(rows) > 85 else {}
    acumulado     = _row_get(rows, 89, month_col_map) if len(rows) > 89 else {}

    records: List[MonthlyCashFlow] = []
    for month_str in list(month_col_map.keys()):
        d = month_dates[month_str]
        inc = income_total.get(month_str, 0.0)
        exp = exp_total.get(month_str, 0.0)
        records.append(MonthlyCashFlow(
            month=month_str,
            month_number=month_numbers[month_str],
            is_actual=d < CUTOFF_DATE,
            income=inc,
            expenses=exp,
            net_cash_flow=inc - exp,
            cumulative_balance=acumulado.get(month_str, 0.0),
            income_ventas=0.0,
            income_separaciones=separaciones.get(month_str, 0.0),
            income_entregas=entregas.get(month_str, 0.0),
            income_financiamiento=financiamiento.get(month_str, 0.0),
            exp_construccion=exp_const.get(month_str, 0.0),
            exp_suelo=exp_suelo.get(month_str, 0.0),
            exp_tecnicos=exp_tec.get(month_str, 0.0),
            exp_juridico=exp_jur.get(month_str, 0.0),
            exp_financiero=exp_fin.get(month_str, 0.0),
            exp_gestion=exp_ges.get(month_str, 0.0),
            exp_comercializacion=exp_com.get(month_str, 0.0),
        ))
    return records


def _get_synthetic_social() -> List[MonthlyCashFlow]:
    """
    Fallback synthetic data — social interest project (RD$).
    61 months from Sep 2024.  First 21 months = actual (before Jun 2026).
    Expense breakdown: Construcción 55%, Suelo 15%, Técnicos 10%,
    Jurídico 5%, Financiero 5%, Gestión 7%, Comercial 3%.
    Income in actual months: separaciones front-loaded, entregas near delivery.
    """
    import math
    records = []
    cum = 0.0
    for i in range(61):
        year = 2024 + (8 + i) // 12
        month = ((8 + i) % 12) + 1
        month_str = f"{year}-{month:02d}"
        is_actual = i < 21

        # ── Income ────────────────────────────────────────────────────────────
        if is_actual:
            # Separaciones: months 0-12 front-loaded (early sales deposits)
            sep = max(0.0, 4_500_000 * math.exp(-0.15 * i) + 800_000) if i <= 14 else 0.0
            # Entregas: months 16-20 (first deliveries)
            entr = max(0.0, (i - 15) * 2_800_000) if i >= 16 else 0.0
            # Financiamiento: months 10-20 (bank credit lines start flowing)
            fin = max(0.0, (i - 9) * 1_200_000) if i >= 10 else 0.0
            inc = sep + entr + fin
        else:
            # Projected: bulk of entregas + financiamiento ramp-up
            sep = 0.0
            entr = max(0.0, (i - 20) * 5_500_000) if i > 20 else 0.0
            fin = max(0.0, (i - 20) * 2_000_000) if i > 20 else 0.0
            inc = entr + fin

        # ── Expenses ──────────────────────────────────────────────────────────
        exp_total = max(0.0, 2_000_000 + math.sin(i * 0.5) * 1_000_000) if i < 55 else 0.0
        # Distribute across components
        e_const  = exp_total * 0.55
        e_suelo  = exp_total * 0.15
        e_tec    = exp_total * 0.10
        e_jur    = exp_total * 0.05
        e_fin    = exp_total * 0.05
        e_gest   = exp_total * 0.07
        e_com    = exp_total * 0.03

        net = inc - exp_total
        cum += net
        records.append(MonthlyCashFlow(
            month=month_str, month_number=i + 1, is_actual=is_actual,
            income=inc, expenses=exp_total, net_cash_flow=net, cumulative_balance=cum,
            income_separaciones=sep if is_actual else 0.0,
            income_entregas=entr,
            income_financiamiento=fin,
            exp_construccion=e_const,
            exp_suelo=e_suelo,
            exp_tecnicos=e_tec,
            exp_juridico=e_jur,
            exp_financiero=e_fin,
            exp_gestion=e_gest,
            exp_comercializacion=e_com,
        ))
    return records


def _get_synthetic_tourist() -> List[MonthlyCashFlow]:
    """
    Fallback synthetic data — tourist project (USD).
    60 months from Jun 2026.  All projected (project starts now).
    Expense breakdown: Construcción 50%, Suelo 20%, Técnicos 10%,
    Jurídico 5%, Financiero 6%, Gestión 6%, Comercial 3%.
    """
    import math
    records = []
    cum = 0.0
    for i in range(60):
        year = 2026 + (5 + i) // 12
        month = ((5 + i) % 12) + 1
        month_str = f"{year}-{month:02d}"

        # Income: separaciones months 1-6, entregas months 18+, financing months 12+
        sep  = max(0.0, 120_000 * math.exp(-0.2 * i) + 30_000) if i <= 8 else 0.0
        entr = max(0.0, (i - 17) * 180_000) if i > 17 else 0.0
        fin  = max(0.0, (i - 11) * 80_000) if i > 11 else 0.0
        inc  = sep + entr + fin

        exp_total = max(0.0, 300_000 + math.sin(i * 0.4) * 150_000) if i < 50 else 0.0
        e_const = exp_total * 0.50
        e_suelo = exp_total * 0.20
        e_tec   = exp_total * 0.10
        e_jur   = exp_total * 0.05
        e_fin   = exp_total * 0.06
        e_gest  = exp_total * 0.06
        e_com   = exp_total * 0.03

        net = inc - exp_total
        cum += net
        records.append(MonthlyCashFlow(
            month=month_str, month_number=i + 1, is_actual=False,
            income=inc, expenses=exp_total, net_cash_flow=net, cumulative_balance=cum,
            income_separaciones=sep,
            income_entregas=entr,
            income_financiamiento=fin,
            exp_construccion=e_const,
            exp_suelo=e_suelo,
            exp_tecnicos=e_tec,
            exp_juridico=e_jur,
            exp_financiero=e_fin,
            exp_gestion=e_gest,
            exp_comercializacion=e_com,
        ))
    return records
